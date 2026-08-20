from flask import Flask, request, render_template
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import os

app = Flask(__name__)


@app.route("/")
def home():
    return render_template("form.html")


@app.route("/submit", methods=["POST"])
def submit():

    name = request.form["name"]
    email = request.form["email"]
    phone = request.form["phone"]
    age = request.form["age"]

    file_name = "submissions.xlsx"

    # Create Excel file if it doesn't exist
    if not os.path.exists(file_name):

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Submissions"

        # Add headers
        sheet.append([
            "ID",
            "Name",
            "Email",
            "Phone",
            "Age",
            "Submission Date & Time"
        ])

        # Format headers
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        workbook.save(file_name)

    # Open existing Excel file
    workbook = load_workbook(file_name)
    sheet = workbook["Submissions"]

    # Generate ID
    new_id = sheet.max_row

    # Current date and time
    submission_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Add new submission
    sheet.append([
        new_id,
        name,
        email,
        phone,
        age,
        submission_time
    ])

    # Set column widths
    sheet.column_dimensions["A"].width = 8
    sheet.column_dimensions["B"].width = 25
    sheet.column_dimensions["C"].width = 30
    sheet.column_dimensions["D"].width = 18
    sheet.column_dimensions["E"].width = 10
    sheet.column_dimensions["F"].width = 25

    # Center some columns
    for row in sheet.iter_rows():
        row[0].alignment = Alignment(horizontal="center")
        row[4].alignment = Alignment(horizontal="center")
        row[5].alignment = Alignment(horizontal="center")

    # Freeze header
    sheet.freeze_panes = "A2"

    # Create Excel table
    if "SubmissionTable" not in sheet.tables:

        table = Table(
            displayName="SubmissionTable",
            ref=f"A1:F{sheet.max_row}"
        )

        table_style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

        table.tableStyleInfo = table_style
        sheet.add_table(table)

    else:
        # Update table range when a new row is added
        table = sheet.tables["SubmissionTable"]
        table.ref = f"A1:F{sheet.max_row}"

    workbook.save(file_name)

    print("Form submitted successfully!")
    print("Name:", name)
    print("Email:", email)
    print("Phone:", phone)
    print("Age:", age)
    print("Submission Time:", submission_time)

    return "Form submitted successfully!"


if __name__ == "__main__":
    app.run(debug=True)